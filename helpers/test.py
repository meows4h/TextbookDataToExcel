import unittest
import os
import random
import pandas as pd
import openpyxl
import datetime
import configparser

import helpers.analytics as alma
import helpers.bookstore as book
import helpers.classes as classes
import helpers.emails as emails
import helpers.enrollment as enroll
import helpers.grabber as outlook
import helpers.gui as gui
import helpers.helpergui as helpgui
import helpers.modes as modes
import helpers.output as output
import helpers.utilities as util

from warnings import simplefilter

# skipping selenium oriented functions as they require logging in via the interface
# this could be done in headless mode w/ more looking over the HTML of the log in
# pages, but focusing on the core logic functions for now

class AnalyticsTest(unittest.TestCase):
    """Testing the analytics.py file."""
    def setUp(self):
        # error message setup
        self.null_msg = "get_columns: Null % SQL Section."
        self.empty_msg = "get_columns: Empty % SQL Section."
        self.missing_msg = "get_columns: Missing Key or Column values in % SQL."
        self.sql_statements = ["SELECT", "FROM", "WHERE", "LIKE"]

    # helper functions

    # checks the SQL columns for proper dictionary setup
    def check_columns(self, columns):
        check = True
        for dict in columns:
            if not dict["Key"] or not dict["Cols"]:
                check = False
                break
        return check
    
    # functionizing the error messages
    def get_sql_msg(self, value):
        new_null = self.null_msg.replace("%", value)
        new_empty = self.empty_msg.replace("%", value)
        new_missing = self.missing_msg.replace("%", value)
        result = {"Null": new_null, "Empty": new_empty, "Miss": new_missing}
        return result

    # testing get_columns
    def test_ebook_sql(self):
        section, columns = alma.get_columns(key="ebook")
        messages = self.get_sql_msg("Ebook")
        
        self.assertIsNotNone(section, messages["Null"])
        self.assertTrue(section, messages["Empty"])

        col_check = self.check_columns(columns)
        self.assertTrue(col_check, messages["Miss"])

    def test_access_sql(self):
        section, columns = alma.get_columns(key="access")
        messages = self.get_sql_msg("Access")
        
        self.assertIsNotNone(section, messages["Null"])
        self.assertTrue(section, messages["Empty"])

        col_check = self.check_columns(columns)
        self.assertTrue(col_check, messages["Miss"])

    def test_physical_sql(self):
        section, columns = alma.get_columns(key="physical")
        messages = self.get_sql_msg("Physical")
        
        self.assertIsNotNone(section, messages["Null"])
        self.assertTrue(section, messages["Empty"])

        col_check = self.check_columns(columns)
        self.assertTrue(col_check, messages["Miss"])
    
    def test_einventory_sql(self):
        section, columns = alma.get_columns(key="e-inventory")
        messages = self.get_sql_msg("E-Inventory")
        
        self.assertIsNotNone(section, messages["Null"])
        self.assertTrue(section, messages["Empty"])

        col_check = self.check_columns(columns)
        self.assertTrue(col_check, messages["Miss"])

    def test_public_sql(self):
        section, columns = alma.get_columns(key="public")
        messages = self.get_sql_msg("Public")
        
        self.assertIsNotNone(section, messages["Null"])
        self.assertTrue(section, messages["Empty"])

        col_check = self.check_columns(columns)
        self.assertTrue(col_check, messages["Miss"])

    # testing get_col_len
    def test_col_len(self):
        # example from real code
        # expects 14 columns
        columns = [
            {
                "Key": "Bibliographic Details",
                "Cols": [
                    "Author",
                    "Title",
                    "ISBN",
                    "Resource Type",
                    "MMS Id",
                    "Earliest Possible Publication Year",
                ],
            },
            {
                "Key": "Physical Item Details",
                "Cols": [
                    "Material Type",
                    "Num of Items (In Repository)",
                    "Temporary Physical Location in Use",
                ],
            },
            {
                "Key": "Holdings Details",
                "Cols": ["Holdings Lifecycle", "Suppressed from Discovery"],
            },
            {"Key": "Edition Simplified", "Cols": ["Edition Simplified (Num)"]},
            {"Key": "Location", "Cols": ["Location Name"]},
            {"Key": "Temporary Location", "Cols": ["Temporary Location Name"]},
        ]
        num = alma.get_col_len(columns)
        expected = 14
        self.assertEqual(num, expected, f"get_col_len: Miscounting column count; expected {expected}, got {num}")

    # selenium functions would go here

    # testing setup_sql
    def test_setup_sql(self):
        col_keys = ["ebook", "access", "physical", "e-inventory", "public"]
        for key_val in col_keys:
            section, columns = alma.get_columns(key=key_val)
            sql = alma.setup_sql(section, columns)
            for command in self.sql_statements:
                self.assertTrue(command in sql, f"setup_sql: {key_val} SQL statement missing {command}")

    # export and import functions would go here


class BookstoreTest(unittest.TestCase):
    """Testing the bookstore.py file."""
    def setUp(self):
        self.clean_msg = "str_clean: Value mismatch from expected; % =/= ^"

    def get_clean_msg(self, res, exp):
        out_msg = self.clean_msg.replace("%", res)
        out_msg = out_msg.replace("^", exp)
        return out_msg

    # testing str_clean
    def test_clean_nl(self):
        new_line = "\nsome text\n\n "
        result = book.str_clean(new_line)
        expected = "some text"
        message = self.get_clean_msg(result, expected)
        self.assertEqual(result, expected, message)

    def test_clean_tab(self):
        tab_char = "\t\thello \t world"
        result = book.str_clean(tab_char)
        expected = "hello world"
        message = self.get_clean_msg(result, expected)
        self.assertEqual(result, expected, message)

    def test_clean_space(self):
        space_text = "   something     a      "
        result = book.str_clean(space_text)
        expected = "something a"
        message = self.get_clean_msg(result, expected)
        self.assertEqual(result, expected, message)

    def test_clean(self):
        text = "  \t\n\n something \t    t  here   \t   "
        result = book.str_clean(text)
        expected = "something t here"
        message = self.get_clean_msg(result, expected)
        self.assertEqual(result, expected, message)

    # get_page_soup selenium function

    # testing get_link

    # get_prices selenium function

    # pull_textbook_data uses selenium

    # pull_info import function
    def test_pull_info(self):
        directory = ["testing", "test_bookstore.csv"]
        path = get_directory(directory)
        table = book.pull_info(path)

        term = "2022-Summer"
        subject = "MTH : Mathematics"
        code = "101"
        section = "001"
        instructor = "LastName, FirstName"
        title = "Intro to Math Textbook"
        edition = "11"
        author = "Textbook Author"
        isbn = "978-1-11-111"
        publisher = "Textbook Publisher"
        req = "Optional"
        sku = "1"
        comments = "Some Comment"
        requisition = "2/2/2022 2:22:22 PM"

        example_row = [term, subject, code, section, instructor, title, edition, author, isbn, publisher, req, sku, comments, requisition]
        row = table[0]
        
        for idx, cell in enumerate(row):
            self.assertEqual(cell, example_row[idx], f"pull_info: {cell} doesn't equal {example_row[idx]}")


class ClassesTest(unittest.TestCase):
    """Testing the classes.py file."""
    def setUp(self):
        # individiual preset testing
        self.analytics_data = {"Physical MMS Id": {
            "Types": ["Physical"],
            "Copies": [2],
            "Users": [0],
            "CDL": [False],
            "Link": "https://www.example.com",
            "Year": 2024,
            "Location": "Valley Library BIB"
        },
        "Ebook MMS Id": {
            "Types": ["Electronic"],
            "Copies": [0],
            "Users": [4],
            "CDL": [True],
            "Link": "https://www.example2.com",
            "Platform": "",
            "Year": 2020
        }}
        self.book_info = {
            "Title": "Book Title",
            "Author": "Book Author",
            "Edition": "Book Edition",
            "Instructor": "Example Name",
            "Email": "Example Email",
            "Course": "Example Course",
            "Section": 1,
            "Enroll": [1, "C"],
            "ISBN": 1,
            "Publisher": "Book Publisher",
            "Req": "Optional",
            "RequiDate": "Requisition Date",
            "Comment": "Example Comment",
            "Analytics": self.analytics_data,
        }
        self.book = classes.Book(self.book_info)

    def generate_book_list(self):
        # randomized testing
        self.book_list = []
        self.course_list = []
        self.section_list = []

        # course_templates = ["MTH", "BI", "WGSS", "CHEM", "GEO"]
        section_templates = [1, 100, 200, 300, 400, 500]
        campus_templates = ["C", "D", "N", "B", "Z", "L", "H", "PDX"]
        platform_templates = ["OverDrive", "EBSCO", "Elsevier", "ProQuest", ""]
        book_count = get_random(5, 100)
        for num in range(1, book_count + 1):
            base_course = f"Course 0"
            base_section = section_templates[get_random(0, len(section_templates) - 1)]
            base_enroll = [get_random(0, 150), campus_templates[get_random(0, len(campus_templates ) - 1)]]
            edition = get_random(0, 20)
            # sometimes the edition value is left empty
            if edition == 0:
                edition = None
            title = f"Book Title {num}"
            author = f"Book Author {num}"
            instructor = f"Example Name {num}"
            email = f"Example Email {num}"
            publisher = f"Book Publisher {num}"
            req_check = get_random(1, 2)
            if req_check == 2:
                requirements = "Required"
            else:
                requirements = "Optional"

            base_analytics = {}
            rand = get_random(1, 3)
            if rand >= 2:
                base_analytics[f"{num}Physical"] = {
                    "Types": ["Physical"],
                    "Copies": [get_random(1, 11)],
                    "Users": [0],
                    "CDL": [False],
                    "Link": "https://www.example3.com",
                    "Year": 2000 + get_random(0, 25),
                    "Location": "Valley Library BIB"
                }
            
            rand = get_random(1, 3)
            if rand >= 2:
                rand = get_random(1, 2)
                if rand >= 2:
                    is_cdl = True
                else:
                    is_cdl = False
                base_platform = platform_templates[get_random(0, len(platform_templates) - 1)]
                base_analytics[f"{num}Electronic"] = {
                    "Types": ["Electronic"],
                    "Copies": [0],
                    "Users": [get_random(1, 11)],
                    "CDL": [is_cdl],
                    "Link": "https://www.example4.com",
                    "Platform": base_platform,
                    "Year": 2000 + get_random(0, 25)
                }

            base_info = {
                "Title": title,
                "Author": author,
                "Edition": edition,
                "Instructor": instructor,
                "Email": email,
                "Course": base_course,
                "Section": base_section,
                "Enroll": base_enroll,
                "ISBN": num,
                "Publisher": publisher,
                "Req": requirements,
                "RequiDate": "2/2/2022 2:22:22 PM",
                "Comment": "Example Comment",
                "Analytics": base_analytics,
            }

            new_book = classes.Book(base_info)

            book_sections = []

            # generating sections for the already existing base section
            section_count = get_random(1, 30)
            for sec in range(1, section_count + 1):
                course_section = section_templates[get_random(0, len(section_templates) - 1)]
                sec_num = course_section + sec
                sec_enroll = [get_random(0, 150), campus_templates[get_random(0, len(campus_templates) - 1)]]
                new_book.add_section(base_course, sec_num, instructor, email, sec_enroll)

            book_sections.append(section_count + 1)

            # generate remaining courses + sections
            course_count = get_random(1, 6)
            for course in range(1, course_count + 1):
                course_name = f"Course {course}"
                course_section = section_templates[get_random(0, len(section_templates) - 1)]
                course_enroll = [get_random(0, 150), campus_templates[get_random(0, len(campus_templates) - 1)]]
                new_book.add_course(course_name, course_section, instructor, email, course_enroll)

                section_count = get_random(1, 30)
                for sec in range(1, section_count + 1):
                    sec_num = course_section + sec
                    sec_enroll = [get_random(0, 150), campus_templates[get_random(0, len(campus_templates ) - 1)]]
                    new_book.add_section(course_name, sec_num, instructor, email, sec_enroll)

                book_sections.append(section_count + 1)
            
            book_sections = sorted(book_sections, reverse=True)
            self.section_list.append(book_sections)
            
            # need to add one for the initial set
            self.course_list.append(course_count + 1)
            self.book_list.append(new_book)

    # Book Class testing

    # add_course method
    def test_add_course(self):
        course_name = "Test Course"
        section_num = 999
        instructor = "Test Name"
        email = "Test Email"
        enroll_data = [99, "D"]

        self.book.add_course(course_name, section_num, instructor, email, enroll_data)
        self.assertEqual(len(self.book.courses), 2, "add_course: Number of courses is incorrect.")
        self.assertEqual(len(self.book.sections), 2, "add_course: Number of sections is incorrect.")
        self.assertEqual(self.book.sec_size[-1], 1, "add_course: Incorrect sec_size array number.")
        self.assertEqual(self.book.total_enroll, 100, "add_course: Total enrollment count is incorrect.")

    # add_section method
    def test_add_section(self):
        course_name = "Example Course"
        section_num = 999
        instructor = "Test Name"
        email = "Test Email"
        enroll_data = [99, "D"]

        self.book.add_section(course_name, section_num, instructor, email, enroll_data)
        self.assertEqual(len(self.book.courses), 1, "add_section: Number of courses is incorrect.")
        # adding the 0 index as book.sections is a list of lists of sections, with the index
        # corresponding to the index of the course index
        self.assertEqual(len(self.book.sections[0]), 2, "add_section: Number of sections is incorrect.")
        self.assertEqual(self.book.sec_size[-1], 2, "add_section: Incorrect sec_size array number.")
        self.assertEqual(self.book.total_enroll, 100, "add_section: Total enrollment count is incorrect.")

    # add_isbn method
    def test_add_isbn(self):
        isbn = 2
        self.book.add_isbn(isbn)
        self.assertEqual(len(self.book.isbns), 2, "add_isbn: Number of ISBNs is incorrect.")
        self.assertIn(2, self.book.isbns, "add_isbn: Missing ISBN number.")

    # add_enroll method
    def test_add_enroll(self):
        enroll_data = [99, "D"]
        self.book.add_enroll(enroll_data[1], enroll_data[0])
        self.assertEqual(self.book.total_enroll, 100, "add_enroll: Total enrollment count is incorrect.")
        self.assertIn("Ecampus", self.book.campuses, "add_enroll: Missing one of the correct campuses.")

    # add_required method
    def test_add_required(self):
        self.book.add_required()
        self.assertEqual("Required", self.book.requirement, "add_required: Value did not update properly.")

    # End Book Class testing

    # get_max_index testing
    def test_max_index(self):
        # setup
        for num in range(1, 5):
            course_name = f"Test Course {num}"
            base_section = 1
            base_enroll = [99, "D"]
            instructor = "Test Name"
            email = "Test Email"
            self.book.add_course(course_name, base_section, instructor, email, base_enroll)
            for sec in range(1, num):
                sec_num = base_section + sec
                sec_enroll = [sec + num, "C"]
                self.book.add_section(course_name, sec_num, instructor, email, sec_enroll)

        max_list = classes.get_max_index(self.book.sec_size)
        # it finds the first index first, meaning the matching sizes are done by lowest index
        expected = [4, 3, 2, 0, 1]
        self.assertEqual(expected, max_list, "get_max_index: Course maximum section count order does not match expected values.")

    # get_max_courses testing
    def test_max_courses(self):
        self.generate_book_list()
        expected_num = max(self.course_list)
        most_courses = classes.get_max_courses(self.book_list)
        self.assertEqual(most_courses, expected_num, "")

    # get_max_sections_list testing
    def test_max_sections_list(self):
        self.generate_book_list()
        expected_size = max(self.course_list)

        most_courses = classes.get_max_courses(self.book_list)
        section_size_list = classes.get_max_sections_list(self.book_list, most_courses)

        max_list_size = 0
        for sec_list in self.section_list:
            max_list_size = max(len(sec_list), max_list_size)

        expected_list = []
        for idx in range(0, max_list_size + 1):
            max_idx = 0
            for sec_list in self.section_list:
                if len(sec_list) - 1 >= idx:
                    max_idx = max(sec_list[idx], max_idx)
            expected_list.append(max_idx)

        for idx, sec_size in enumerate(section_size_list):
            self.assertGreater(expected_size, idx)
            self.assertEqual(sec_size, expected_list[idx])

    # process_book testing
    def test_process_book(self):
        self.generate_book_list()
        local_list = classes.process_book(self.book_list, self.book_info)
        self.book_list.append(self.book)
        for book_idx in range(0, len(local_list)):
            self.assertEqual(local_list[book_idx], self.book_list[book_idx])

    # process_courses testing
    def test_process_courses(self):
        simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
        simplefilter(action="ignore", category=FutureWarning)

        self.generate_book_list()
        dataframe = pd.DataFrame()
        head_names, main_headers = util.get_sheet_headers()
        total_cols = 0
        for header in main_headers:
            dataframe[f"{header}"] = []
            total_cols += 1
        format_headers = util.get_format_headers()

        max_len = 0
        for book in self.book_list:
            max_len = max(len(book.courses), max_len)
        total_cols += max_len

        max_courses = classes.process_courses(self.book_list, format_headers, dataframe)
        self.assertEqual(max_len, max_courses)
        self.assertEqual(total_cols, len(dataframe.columns))

    # process_sections testing
    def test_process_sections(self):
        simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
        simplefilter(action="ignore", category=FutureWarning)

        self.generate_book_list()
        dataframe = pd.DataFrame()
        head_names, main_headers = util.get_sheet_headers()
        total_cols = 0
        for header in main_headers:
            dataframe[f"{header}"] = []
            total_cols += 1
        format_headers = util.get_format_headers()

        max_len = 0
        for book in self.book_list:
            max_len = max(len(book.courses), max_len)
        total_cols += max_len

        most_courses = classes.get_max_courses(self.book_list)
        section_size_list = classes.get_max_sections_list(self.book_list, most_courses)
        for sec_size in section_size_list:
            total_cols += sec_size * 4

        max_courses = classes.process_courses(self.book_list, format_headers, dataframe)
        classes.process_sections(self.book_list, format_headers, max_courses, dataframe)

        self.assertEqual(total_cols, len(dataframe.columns))

    # process_isbns testing
    def test_process_isbns(self):
        simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
        simplefilter(action="ignore", category=FutureWarning)

        self.generate_book_list()
        dataframe = pd.DataFrame()
        head_names, main_headers = util.get_sheet_headers()
        total_cols = 0
        for header in main_headers:
            dataframe[f"{header}"] = []
            total_cols += 1
        format_headers = util.get_format_headers()

        max_len = 0
        max_isbn = 0
        for book in self.book_list:
            max_len = max(len(book.courses), max_len)
            max_isbn = max(len(book.isbns), max_isbn)
        total_cols += max_len

        most_courses = classes.get_max_courses(self.book_list)
        section_size_list = classes.get_max_sections_list(self.book_list, most_courses)
        for sec_size in section_size_list:
            total_cols += sec_size * 4

        max_courses = classes.process_courses(self.book_list, format_headers, dataframe)
        classes.process_sections(self.book_list, format_headers, max_courses, dataframe)
        classes.process_isbns(self.book_list, head_names, main_headers, dataframe)

        self.assertEqual(total_cols, len(dataframe.columns))

    # import_data testing
    def test_import_data(self):
        simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
        simplefilter(action="ignore", category=FutureWarning)

        self.generate_book_list()
        dataframe = pd.DataFrame()
        head_names, main_headers = util.get_sheet_headers()
        for header in main_headers:
            dataframe[f"{header}"] = []
            dataframe[f"{header}"] = dataframe[f"{header}"].astype(object)
        format_headers = util.get_format_headers()

        max_courses = classes.process_courses(self.book_list, format_headers, dataframe)
        classes.process_sections(self.book_list, format_headers, max_courses, dataframe)
        classes.process_isbns(self.book_list, head_names, main_headers, dataframe)
        dataframe = classes.import_data(self.book_list, format_headers, head_names, dataframe)
        self.assertEqual(len(self.book_list), len(dataframe))


class EmailsTest(unittest.TestCase):
    """Testing the emails.py file."""
    def setUp(self):
        book_title = "Book Title"
        book_author = "Book Author"
        book_edition = "Book Edition"
        book_year = "Book Year"
        book_access = {
            "Ebook": {
                "Number": 1,
                "Link": "ebook cdl link",
                "CDL": True
            },
            "Print1": {
                "Number": 2,
                "Link": "print 1 link"
            },
            "Print2": {
                "Number": 4,
                "Link": "print 2 link"
            },
            "Audio": {
                "Link": "audio link"
            }
        }
        self.emails_book = emails.Book(book_title, book_author, book_edition, book_year, book_access)

        inst_name = "Instructor Name"
        inst_email = "Instructor Email"
        inst_course = "Course 1"
        inst_section = "101"
        self.emails_instructor = emails.Instructor(inst_name, inst_email, inst_course, inst_section, self.emails_book)

    # Instructor Class testing

    # add_book method
    def test_add_book(self):
        title = "Book 2"
        author = "Author 2"
        edition = "Edition 2"
        year = "Year 2"
        access = {"Ebook": {
                "Number": 7,
                "Link": "ebook link",
                "CDL": False
                }}
        local_book = emails.Book(title, author, edition, year, access)

        case_courses = ["Course 1", "Course 1", "Course 2"]
        case_sections = ["101", "102", "101"]
        for idx, course in enumerate(case_courses):
            self.emails_instructor.add_book(course, case_sections[idx], local_book)
                    
        self.assertEqual(2, len(self.emails_instructor.data))
        self.assertEqual(2, len(self.emails_instructor.data["Course 1"]))
        self.assertEqual(2, len(self.emails_instructor.data["Course 1"]["101"]))
        self.assertEqual(1, len(self.emails_instructor.data["Course 1"]["102"]))
        self.assertEqual(1, len(self.emails_instructor.data["Course 2"]))
        self.assertEqual(1, len(self.emails_instructor.data["Course 2"]["101"]))

        # could add in more tests beyond just the length checking but the data issues would
        # be apparent when generating emails as well
        

    # End Instructor Class testing

    # update_excel testing
    # this needs a setup to run sheetmaker and import other information?
    def test_update_excel(self):
        bookstore_dir = get_directory(["testing", "test_bookstore.csv"])
        output_dir = get_directory(["testing", "test_output.xlsx"])
        
        workbook = openpyxl.load_workbook(output_dir)
        sheet_name = "Test"
        worksheet = workbook[sheet_name]

        header_dict, header_list = util.get_sheet_headers()

        # generate test data
        email_sent_list = []
        email_ready_list = []
        test_time = datetime.datetime.now()
        for idx, row in enumerate(worksheet.iter_rows(min_row=2)):
            if idx >= 151:
                break
            prior_check = False
            for cell in row:
                col = cell.column_letter
                if worksheet[f"{col}1"].value == header_dict["ReadingList"]:
                    num = get_random(1, 2)
                    if num == 1:
                        cell.value = True
                        prior_check = True
                        email_ready_list.append(idx)
                    else:
                        cell.value = False
                elif worksheet[f"{col}1"].value == header_dict["EmailDate"]:
                    num = get_random(1, 2)
                    if num == 1 or prior_check == False:
                        cell.value = None
                    else:
                        email_ready_list.pop()
                        email_sent_list.append(idx)
                        cell.value = test_time

        workbook.save(output_dir)
        workbook.close()

        emails.update_excel(output_dir, email_ready_list, sheet_name)

        final_email_list = sorted(email_ready_list + email_sent_list)

        workbook = openpyxl.load_workbook(output_dir)
        worksheet = workbook[sheet_name]
        for idx, row in enumerate(worksheet.iter_rows(min_row=2)):
            if idx >= 151:
                break
            reading_val = None
            email_val = None
            for cell in row:
                col = cell.column_letter
                if worksheet[f"{col}1"].value == header_dict["ReadingList"]:
                    reading_val = cell.value
                elif worksheet[f"{col}1"].value == header_dict["EmailDate"]:
                    email_val = cell.value
            if idx in final_email_list:
                self.assertTrue(reading_val)
                self.assertTrue(email_val)
            else:
                self.assertFalse(reading_val)
                self.assertFalse(email_val)


    # create_email_excel testing
    def test_create_emails(self):
        output_dir = get_directory(["testing", "test_output.xlsx"])
        
        workbook = openpyxl.load_workbook(output_dir)
        sheet_name = "Test"
        worksheet = workbook[sheet_name]

        header_dict, header_list = util.get_sheet_headers()

        # generate test data
        email_sent_list = []
        email_ready_list = []
        test_time = datetime.datetime.now()
        for idx, row in enumerate(worksheet.iter_rows(min_row=2)):
            if idx >= 151:
                break
            prior_check = False
            for cell in row:
                col = cell.column_letter
                if worksheet[f"{col}1"].value == header_dict["ReadingList"]:
                    num = get_random(1, 2)
                    if num == 1:
                        cell.value = True
                        prior_check = True
                        email_ready_list.append(idx)
                    else:
                        cell.value = False
                elif worksheet[f"{col}1"].value == header_dict["EmailDate"]:
                    num = get_random(1, 2)
                    if num == 1 or prior_check == False:
                        cell.value = None
                    else:
                        email_ready_list.pop()
                        email_sent_list.append(idx)
                        cell.value = test_time

        workbook.save(output_dir)
        workbook.close()
        emails.create_email_excel("Test", output_dir)
        
        final_email_list = sorted(email_ready_list + email_sent_list)
        email_dir = get_directory(["..", "email_output.xlsx"])

        email_workbook = openpyxl.load_workbook(email_dir)
        email_worksheet = email_workbook["Email List"]

        for row in email_worksheet.iter_rows(min_row=2):
            for cell in row:
                col = cell.column_letter
                if email_worksheet[f"{col}1"] == "Name":
                    name_col_list = cell.value.split(" ")
                    number = int(name_col_list[2])
                    self.assertIn(number, final_email_list)
                    break

    # write_to_excel testing -- this is an export function and should be fine


class EnrollTest(unittest.TestCase):
    """Testing the enrollment.py file."""
    def setUp(self):
        directory = ["testing", "test_enrollment.csv"]
        self.enroll_dir = get_directory(directory)
        self.instructor_dict = {"Name1 Name2": "Name@email.com"}

    def check_instructor_dict(self, instr_dict, name, email):
        if name not in instr_dict:
            return False
        if email not in instr_dict[name]:
            return False
        return True

    def test_get_enrollment_without(self):
        instr_dict, enroll_dict = enroll.get_enrollment_data(self.enroll_dir)
        self.assertIn("MTH101", enroll_dict, "")
        self.assertIn("1", enroll_dict["MTH101"], "")

        ex_name = "LastName, FirstName"
        ex_email = "FirstName.LastName@example.edu"
        check = self.check_instructor_dict(instr_dict, ex_name, ex_email)
        self.assertTrue(check, "")

    def test_get_enrollment_with(self):
        instr_dict, enroll_dict = enroll.get_enrollment_data(self.enroll_dir, self.instructor_dict)
        self.assertIn("MTH101", enroll_dict, "")
        self.assertIn("1", enroll_dict["MTH101"], "")

        ex_name = "LastName, FirstName"
        ex_email = "FirstName.LastName@example.edu"
        check = self.check_instructor_dict(instr_dict, ex_name, ex_email)
        self.assertTrue(check, "")

        for name in self.instructor_dict:
            email = self.instructor_dict[name]
            check = self.check_instructor_dict(self.instructor_dict, name, email)
            self.assertTrue(check, "")
        

class OutlookTest(unittest.TestCase):
    """Testing the grabber.py file."""
    # every function aside from the importer and exporter deal with selenium webdrivers


class GUITest(unittest.TestCase):
    """Testing the gui.py file."""
    # automated gui testing falls under a similar circumstance to testing selenium
    # there are ways and methods to do so, but for now, not enough time


class HelperGUITest(unittest.TestCase):
    """Testing the helpergui.py file."""
    # see above GUI notes, automated testing gui not within scope as of right now
    # additionally, it is easy to manually test and verify results


class ModesTest(unittest.TestCase):
    """Testing the modes.py file."""
    # functions to primarily test here would be the excel sheet updaters
    
    # enrollment_update testing

    # analytics_update testing
    
    # emails_update testing


class OutputTest(unittest.TestCase):
    """Testing the output.py file."""
    # this file hosts a single function for the pure reason of handling
    # formatting and finishing out the export for an excel file
    # it should be fine to leave out of testing for now


class SheetTest(unittest.TestCase):
    """Testing the sheetmaker.py file."""
    # only testing the version of the function with no gui, no cli interaction

    # make_excel_sheet testing


class UtilTest(unittest.TestCase):
    """Testing the utilities.py file."""

    # get_int testing
    def test_get_int(self):
        values = [" 12   4 5  ", "93478", None, "", "     ", 65.435343, 109]
        expected = [1245, 93478, None, None, None, 65, 109]
        for idx, value in enumerate(values):
            result = util.get_int(value)
            self.assertEqual(expected[idx], result, f"get_int: result {result} does not match {expected[idx]}.")

    # get_clean testing
    def test_get_clean(self):
        test_cleaner = ["Example", "WORDS", "something"]
        test_names = [
            "a book that has examples and wOrdS SOMEthing to do with an example e-reading",
            "what other EXAMPLE case isn't here",
            "'what about words appearing' or something elsewhere"
            ]
        expected = [
            "A Book That Has S And To Do With An E-Reading",
            "What Other Case Isn't Here",
            "'What About Appearing' Or Elsewhere"
            ]
        for idx, name in enumerate(test_names):
            result = util.get_clean(test_cleaner, name)
            self.assertEqual(expected[idx], result)

    # get_state testing
    def test_get_state(self):
        cases = ["true", "hmmm", "anything", "false", "FALSE", "False", "True"]
        expected = [True, True, True, True, True, False, True]
        for idx, case in enumerate(cases):
            result = bool(util.get_state(case))
            self.assertEqual(expected[idx], result)

    # get_directory testing
    def test_get_directory(self):
        test_config = configparser.ConfigParser()
        test_config.read("helpers/testing/test_config.ini")
        test_config = test_config["test"]
        result = util.get_directory("Test", test_config)
        expected = get_directory(["testing", "test_config.csv"])
        self.assertEqual(expected, result, f"get_directory: result {result} does not match {expected}.")

    # get_filepath testing
    def test_get_filepath(self):
        cases = ["testing"]
        expected = [get_directory(["..", "testing"])]
        false_expected = [get_directory(["testing"])]
        for idx, case in enumerate(cases):
            result = util.get_filepath(case)
            false_result = util.get_filepath(case, home=False)
            self.assertEqual(expected[idx], result)
            self.assertEqual(false_expected[idx], false_result)

    # get_letter testing
    def test_get_letter(self):
        cases = [1, 27, 4, 53, 55]
        expected = ["A", "AA", "D", "BA", "BC"]
        for idx, case in enumerate(cases):
            result = util.get_letter(case)
            self.assertEqual(expected[idx], result)

    # get_edition_string testing
    def test_get_edition_string(self):
        cases = [None, "", "5", 29, 10, "1", 11, 21, 22, 13]
        expected = [None, None, "5th", "29th", "10th", "1st", "11th", "21st", "22nd", "13th"]
        for idx, case in enumerate(cases):
            result = util.get_edition_string(case)
            self.assertEqual(expected[idx], result)

    # get_format_headers just grabs a set of information
    # as it is supposed to be the central point of modification for those strings

    # get_replace_header testing
    def test_get_replace_header(self):
        format_headers = util.get_format_headers()
        cases = [["TestNum", "TestSection", "TestEmailInfo"]]
        for idx, case in enumerate(cases):
            result = util.get_replace_header(format_headers[2], case[0], case[1], case[2])
            for val in case:
                self.assertIn(val, result)

    # get_split_course testing
    def test_get_split_course(self):
        cases = ["MTH251", "BIO305", "WR221Z"]
        expected = [["MTH", "251"], ["BIO", "305"], ["WR", "221Z"]]
        for idx, case in enumerate(cases):
            result = util.get_split_course(case)
            self.assertEqual(expected[idx], result)

    # skipping get_input and get_enabled as it is an input oriented function

    # get_sheet_headers, get_config_headers, and get_string_cleaners 
    # are all also is information grabbing functions

    # get_row_info testing
    def test_get_row_info(self):
        # an implementation can be made for row data to be tested, but there is
        # only one spot this gets used in, so it is not relevant to time to get done
        pass

    # get_campus testing
    def test_get_campus(self):
        # this is breakable via any string starting with D and then
        # filling the rest with non-existent values, but this usually does
        # not occur in the natural data this script will be managing
        cases = ["C", "D", "Z", "L", "N", "B", "H", "PDX", "DC", "23432", "CR"]
        expected = ["Corvallis", "Ecampus", "International", "LaGrande", "Newport", "Cascades", "Portland", "Portland", "Ecampus", None, None]
        for idx, case in enumerate(cases):
            result = util.get_campus(case)
            self.assertEqual(expected[idx], result)

    # skipping set_col_format as it is likely to fluctuate and change

# some helper functions to carry out testing various features

def get_directory(folder_list):
    curr_dir = os.path.dirname(__file__)
    for file in folder_list:
        curr_dir = os.path.join(curr_dir, file)
    return curr_dir


def get_random(low, high):
    return random.randint(low, high)
