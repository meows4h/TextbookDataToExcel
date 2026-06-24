import configparser
import pandas as pd
import openpyxl
import threading

from helpers.bookstore import pull_textbook_data, pull_info
from helpers.utilities import get_int, get_state, get_directory
from helpers.utilities import get_format_headers, get_input
from helpers.utilities import get_sheet_headers, get_filepath
from helpers.enrollment import get_enrollment_data
from helpers.grabber import setup_grabber, get_email, grabber_gui
from helpers.grabber import email_exporter, email_importer
from helpers.analytics import setup_analytics
from helpers.analytics import export_analytics, import_analytics
from helpers.analytics import process_analytics
from helpers.emails import create_email_excel

# this whole file PROBABLY could be made more or less obsolete with some refactoring!
# basically just hardcoded some values to always run with, or certain flags rather than pulling from config
# most of these are either similar, exact, or slimmed down versions of their original workflows, just put into a single function
# either way, some (or all) of this code is likely redundant. either other places can use this code, or this code can
# use functions from elsewhere.
# TODO
# refactor this file (or the other parts of the project) to optimize
# project size & functionality


def start_mode(flag):
    """Takes an integer input to start an alternative script method."""
    if flag == 2:
        csv_mode()
    if flag == 3:
        email_mode()
    if flag == 4:
        update_mode()


def csv_mode():
    """Handles all the options related to updating the CSV files."""
    full_config = configparser.ConfigParser()
    full_config.read("config.ini")

    print("Select CSV mode:")
    text = """1. Email Data
2. Bookstore Data
3. Analytics Data
Input: """
    csv_type = get_input(min=1, max=3, text=text)
    bookstore_cfg = full_config["Textbook"]
    textbk_path = get_directory("Save", bookstore_cfg)

    if csv_type == 1:
        emails_csv(full_config, textbk_path)

    elif csv_type == 2:
        pull_textbook_data()

    elif csv_type == 3:
        analytics_csv(full_config, textbk_path)


def email_mode():
    """Handles creating a new email Excel sheet."""
    create_email_excel()


def update_mode():
    """Handles all the options related to updating the main Excel sheet."""
    # changing sheetname value to manual input
    # full_config = configparser.ConfigParser()
    # full_config.read("config.ini")
    # sheet_name = full_config["Email"]["Sheetname"]
    sheet_name = get_input(text="Input which sheet is being updated (i.e. Spring26): ")

    print("Select Update mode:")
    text = """1. Enrollment Data
2. Analytics Data
3. Email Inputs
Input: """
    update_type = get_input(min=1, max=3, text=text)
    if update_type == 1:
        enrollment_update(sheet_name)
    elif update_type == 2:
        analytics_update(sheet_name)
    elif update_type == 3:
        emails_update(sheet_name)


def emails_csv(full_config, textbk_path, pre_import=None):
    """Updates the email csv file with updated information from Outlook."""
    grabber_cfg = full_config["Grabber"]
    emails_path = get_directory("Save", grabber_cfg)

    if pre_import == None:
        import_data = get_import()
    else:
        import_data = pre_import

    if import_data:
        try:
            emails = email_importer(emails_path)
        except BaseException:
            emails = {}
    else:
        emails = {}

    textbook_table = pull_info(textbk_path)

    # this is the CLI implementation
    if pre_import == None:
        # email implementation
        grabber_driver = setup_grabber()

        for row in textbook_table:
            instructor = row[4]
            if instructor in emails:
                email = emails[f"{instructor}"]
            else:
                email = get_email(instructor, grabber_driver)
                emails[f'{instructor}'] = email
                print(email)
            # exporting after each case to ensure that crashes doesnt lead to data
            # loss
            email_exporter(emails_path, emails)
        
        email_exporter(emails_path, emails)
    
    # GUI implementation
    else:
        gui_thread = threading.Thread(target=lambda: grabber_gui(textbook_table, emails))
        gui_thread.start()


def analytics_csv(full_config, textbk_path, pre_import=None):
    """Updates the analytics csv file with current data from Alma analytics."""
    analytics_cfg = full_config["Alma"]
    analytics_path = get_directory("Save", analytics_cfg)

    if pre_import == None:
        import_data = get_import()
    else:
        import_data = pre_import

    if import_data:
        try:
            analytics_store = import_analytics(analytics_path)
        except BaseException:
            analytics_store = {}
    else:
        analytics_store = {}

    # analytics implementation
    textbook_table = pull_info(textbk_path)
    if pre_import == None:
        analytics_driver = setup_analytics()
    else:
        analytics_driver = setup_analytics(gui=True)

    for row in textbook_table:
        isbn = row[8].replace("-", "").strip()
        if isbn == "":
            isbn = None

        if f"{isbn}" in analytics_store:
            continue

        if not (isbn is None or pd.isna(isbn)):
            data = process_analytics(analytics_driver, isbn)

            analytics_store[f"{isbn}"] = {"Data": data}

        # with how long this process takes and because I am running out of time
        # to work out better implementations (i.e. try / except error checking)
        # just exporting the data every row for now
        # TODO
        # update this later with proper error handling to export once
        export_analytics(analytics_path, analytics_store)

    export_analytics(analytics_path, analytics_store)


# TODO
def enrollment_update(sheet_name, file_name=""):
    """Updates the main excel sheet with the data from the enrollment csv file."""
    full_config = configparser.ConfigParser()
    full_config.read("config.ini")
    enroll_config = full_config["Enroll"]
    main_config = full_config["Main"]

    enroll_dir = get_directory("Input", enroll_config)

    if file_name == "":
        directory = get_directory("Output", main_config)
    else:
        directory = get_filepath(file_name)

    header_dict, header_list = get_sheet_headers()
    header_format = get_format_headers()
    instructor_dict, enroll_dict = get_enrollment_data(enroll_dir)
    workbook = openpyxl.load_workbook(directory)
    worksheet = workbook[sheet_name]

    for row in worksheet.iter_rows(min_row=2):
        course_num = 1
        for cell in row:
            col = cell.column_letter
            col_header = worksheet[f"{col}1"].value
            course_header = header_format[0].replace("$", f"{course_num}")
            if col_header == course_header:
                course_num += 1

                # the idea here would be to compare the course code to enrollment data
                # then update the subsequent enrollment section


# this function is functional, but not complete
def analytics_update(sheet_name, file_name=""):
    """Updates the excel sheet with data from the analytics csv file.
    NOTE: It will **NOT** overwrite existing data due to protecting manual data entry.
    """
    full_config = configparser.ConfigParser()
    full_config.read("config.ini")
    alma_config = full_config["Alma"]
    main_config = full_config["Main"]

    alma_dir = get_directory("Save", alma_config)

    if file_name == "":
        directory = get_directory("Output", main_config)
    else:
        directory = get_filepath(file_name)

    header_dict, header_list = get_sheet_headers()
    analytics = import_analytics(alma_dir)
    workbook = openpyxl.load_workbook(directory)
    worksheet = workbook[sheet_name]

    for row in worksheet.iter_rows(min_row=2):
        idx = 1  # this is used to count ISBN columns
        isbn_list = []
        print_list = []
        electronic_list = []

        phys_count = 0
        elec_count = 0
        for cell in row:
            col = cell.column_letter
            isbn_header = header_dict["ISBN"].replace("1", f"{idx}")

            # each time we find an ISBN header
            col_header = worksheet[f"{col}1"].value

            print(
                f"header: {col_header}, cell: {
                    cell.value}, isbn: {isbn_list}, print: {print_list}, elec: {electronic_list}"
            )

            if col_header == isbn_header:
                idx += 1
                isbn = cell.value
                if isbn is not None:
                    isbn = f"{isbn}"
                    isbn_list.append(isbn)

                # check if it is in the data
                if isbn in analytics:
                    data = analytics[isbn]["Data"]
                    if data == {}:
                        continue

                    # create data types to iterate on later in the same row
                    for mms_id in data:
                        for jdx, type in enumerate(data[mms_id]["Types"]):
                            access = type
                            copies = data[mms_id]["Copies"][jdx]
                            users = data[mms_id]["Users"][jdx]
                            cdl = data[mms_id]["CDL"][jdx]
                            link = data[mms_id]["Link"]
                            year = data[mms_id]["Year"]

                            if access == "Physical" and phys_count < 2:
                                phys_count += 1
                                temp = {
                                    "Type": access,
                                    "Number": copies,
                                    "Link": link,
                                    "Year": year,
                                    "MMS": mms_id,
                                }
                                print_list.append(temp)

                            elif access == "Electronic" and elec_count < 1:
                                elec_count += 1
                                temp = {
                                    "Type": access,
                                    "Number": users,
                                    "Link": link,
                                    "Year": year,
                                    "CDL": cdl,
                                    "MMS": mms_id,
                                }
                                electronic_list.append(temp)

            # don't want to modify or overwrite data
            # exception needs to be made for updating CDL value?
            elif cell.value == "" or cell.value is None:
                # this could all be refactored using the keys more than anything else
                # at least i believe
                if phys_count >= 1:
                    if col_header == header_dict["PrintLink1"]:
                        cell.value = print_list[0]["Link"]

                    elif col_header == header_dict["PrintMMSId1"]:
                        cell.value = print_list[0]["MMS"]

                    elif col_header == header_dict["PrintCopies1"]:
                        cell.value = print_list[0]["Number"]

                    elif phys_count >= 2:
                        if col_header == header_dict["PrintLink2"]:
                            cell.value = print_list[1]["Link"]

                        elif col_header == header_dict["PrintMMSId2"]:
                            cell.value = print_list[1]["MMS"]

                        elif col_header == header_dict["PrintCopies2"]:
                            cell.value = print_list[1]["Number"]

                if elec_count >= 1:
                    if col_header == header_dict["EbookLink"]:
                        cell.value = electronic_list[0]["Link"]

                    elif col_header == header_dict["EbookMMSId"]:
                        cell.value = electronic_list[0]["MMS"]

                    elif col_header == header_dict["EbookUsers"]:
                        cell.value = electronic_list[0]["Number"]

            elif elec_count >= 1:
                if col_header == header_dict["IsCDL"]:
                    cell.value = electronic_list[0]["CDL"]

    workbook.save(directory)


def emails_update(sheet_name, file_name=""):
    """Updating excel sheet with data from email csv file."""
    full_config = configparser.ConfigParser()
    full_config.read("config.ini")
    grabber_config = full_config["Grabber"]
    main_config = full_config["Main"]

    email_dir = get_directory("Save", grabber_config)

    if file_name == "":
        directory = get_directory("Output", main_config)
    else:
        directory = get_filepath(file_name)

    header_dict, header_list = get_sheet_headers()
    email_dict = email_importer(email_dir)
    workbook = openpyxl.load_workbook(directory)
    worksheet = workbook[sheet_name]

    # currently what this does:
    # iterate through every cell
    # while doing this, check each prof name that is found, check it
    # against the cache data

    # TODO: this could also:
    # if unfound, search for it in outlook
    # after finding it, catalog it, save it to cache
    # this function would be helpful if STAFF listings are updated after the fact
    # include a y/n option for opening outlook to save time

    for row in worksheet.iter_rows(min_row=2):
        check_next = False
        next_instructor = ""
        for cell in row:
            col = cell.column_letter
            col_header = worksheet[f"{col}1"].value
            if header_dict["Title"] in col_header:
                # breaking this because this is now out of the instructor
                # information
                break

            if check_next:
                if cell.value == "" or pd.isna(cell.value):
                    cell.value = email_dict[next_instructor]
                check_next = False
                next_instructor = ""

            if "Instructor" in col_header:
                if cell.value in email_dict:
                    check_next = True
                    next_instructor = cell.value
                continue

    workbook.save(directory)


def get_import():
    """Basic function to get y/n input for importing data."""
    text = """Import previous data?
Input (y/n): """
    while True:
        try:
            str_input = get_input(text=text)
            if not (str_input == "y" or str_input == "n"):
                raise ValueError("Incorrect input.")
            break
        except Exception as err:
            print(err)

    if str_input == "y":
        importing = True
    elif str_input == "n":
        importing = False
    return importing
