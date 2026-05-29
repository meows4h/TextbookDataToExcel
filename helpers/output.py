from helpers.utilities import get_int, get_letter, get_edition_string
from helpers.utilities import set_col_format
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font

# from openpyxl.styles import Border, Side, Protection


def write_to_sheet(df, sheetname, workbook, headers):
    """Writes the given data to the given Excel file directory.
    Performs stylization functions as well."""
    worksheet = workbook.active
    worksheet.title = sheetname

    row_num = len(df.index)
    col_num = len(df.columns)

    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_letter(col_num)}{row_num}"
    worksheet.auto_filter.add_sort_condition(f"A2:A{row_num}")

    base_start = 0

    yellow_headers = [headers["MaxEnroll"]]
    green_headers = [
        headers["Purchaseable"],
        headers["PurchaseEbook"],
        headers["ReplaceNonPerm"],
        headers["PurchaseCDL"],
        headers["PurchasePrint"],
        headers["PurchaseAudio"],
    ]
    blue_headers = [
        headers["AvailableCatalog"],
        headers["EbookLink"],
        headers["EbookUsers"],
        headers["EbookMMSId"],
        headers["EbookPlatform"],
        headers["IsCDL"],
        headers["PrintLink1"],
        headers["PrintMMSId1"],
        headers["PrintCopies1"],
        headers["PrintLink2"],
        headers["PrintMMSId2"],
        headers["PrintCopies2"],
        headers["AudioLink"],
        headers["AudioMMSId"],
        headers["BNCCopies"],
    ]
    violet_headers = [headers["ReadingList"]]
    purple_headers = [headers["EmailDate"], headers["Notes"]]
    # everything else presumed orange

    for idx, cell in enumerate(worksheet[1]):
        if cell.value in yellow_headers:
            cell.font = Font(bold=True, color="FF000000")
            cell.fill = PatternFill(
                fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00"
            )
        elif cell.value in green_headers:
            cell.font = Font(bold=True, color="FF000000")
            cell.fill = PatternFill(
                fill_type="solid", start_color="FF92D050", end_color="FF92D050"
            )
        elif cell.value in blue_headers:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill(
                fill_type="solid", start_color="FF145F82", end_color="FF145F82"
            )
        elif cell.value in violet_headers:
            cell.font = Font(bold=True, color="FF000000")
            cell.fill = PatternFill(
                fill_type="solid", start_color="FFCA8AFF", end_color="FFCA8AFF"
            )
        elif cell.value in purple_headers:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill(
                fill_type="solid", start_color="FF7030A0", end_color="FF7030A0"
            )
        else:
            cell.font = Font(bold=True, color="FF000000")
            cell.fill = PatternFill(
                fill_type="solid", start_color="FFFFC000", end_color="FFFFC000"
            )
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # grabbing idx when the main headers start
        if cell.value in headers["Title"]:
            base_start = idx

    for col in range(1, col_num + 1):

        col_letter = get_letter(col)

        # setting the base information to be the same size
        if col - 1 < base_start:
            worksheet.column_dimensions[col_letter].width = 12

        # fixing edition numbers
        elif worksheet[col_letter][0].value == headers["Edition"]:
            for cell in worksheet[col_letter]:
                if cell.value == headers["Edition"]:
                    continue
                cell.value = get_edition_string(get_int(cell.value))
            set_col_format(col_letter, False, worksheet)

        # isbn formatting
        elif (
            worksheet[col_letter][0].value.split(" ")[0]
            == headers["ISBN"].split(" ")[0]
        ):
            set_col_format(col_letter, True, worksheet)

        # otherwise not listed
        else:
            set_col_format(col_letter, False, worksheet)

    worksheet.row_dimensions[1].height = 56
